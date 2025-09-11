#Limpeza de dados

from openpyxl import load_workbook
import pandas as pd

df = pd.read_excel('aula_pandas/Boletim.xlsx')
wb = load_workbook('aula_pandas/Boletim.xlsx')
ws = wb.active
ws["A3"] = 'Joana Figueiredo'

wb.save('aula_pandas/Boletim_atualizados.xlsx')

print(wb)