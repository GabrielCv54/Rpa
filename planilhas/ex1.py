##Autromação em planilhas do excel
#Exemplo - planilha de cálculo de média

"""
Biblioteca pandas
Biblioteca openpyxl
"""
import pandas as  pd
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter

#1.Criação do Dataframe com os dados a serem utilizados
data = {
    'Nome':['Leonardo','Matias','Helen','Bianca','Flávio'],
    'Nota 1':[10, 5.6, 7.8, 8.5, 6],
    'Nota 2':[8.6, 7.8, 7.5, 9, 6.4]
}

df= pd.DataFrame(data)

#2- Cálculo da média usando a biblioteca pandas
df['Média'] = df[['Nota 1','Nota 2']].mean(axis=1)
'''print(df)
'''

#3- Salvar em excel
excel_file = 'notas.xlsx'
df.to_excel(excel_file,index=False,sheet_name='Notas')


#4- Abrir a planilha com openpyxl para formatação
wb = load_workbook(excel_file)
ws = wb['Notas']

#5 - Formatação do cabeçalho - Negrito + preenchimento cinza
header_fill = PatternFill(start_color='DDDDDD',fill_type='solid')
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.fill = header_fill

#6 - Formatação das colunas númericas com 2 casas decimais
for row in ws.iter_rows(min_row=2,min_col=2,max_col=5):
   for cell in row:
       cell.number_format = '0.00'

#7. Destacar notas abaixo de 7 com fundo vermelho
for row in ws.iter_rows(min_row=2,min_col=2,max_col=5):
    for cell in row:
        if isinstance(cell.value,(int,float) and cell.value < 7.0):
            cell.fill = PatternFill(start_color='FFCCCC',fill_type='solid')

#8- Salvar as alterações
wb.save(excel_file)

print(f'Planilha: {excel_file} formatada com sucesso')