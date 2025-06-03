import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
pd.options.display.max_rows=(50)
#1 - criar e salvar o Dataframe em um Excel
data = {
    "Produtos":["Computador","Playstation 5","PC",'Impressora Hp deskjet','TV samsung'],
    "Preço":[1140,3500,2960,1520,30500]

}

df = pd.DataFrame(data)
#print(df)


#2- Salvando ele em um excel
df.to_excel('produtos.xlsx',index=False, engine='openpyxl')


#3- Abrir o arquivo Excel com openpyxl para aplicar a formatação
wb = load_workbook('produtos.xlsx') #Abre e carrega o arquivo
ws =wb.active #seleciona a planilha ativa


#4-Deixar o cabeçaho em negrito
for cell in ws[1]:
    cell.font = Font(bold=True,color='4169E1')


#5- Salvar o arquivo
wb.save('produtos_formatados.xlsx')

print('Criação concluída !!')

print(df.to_string())